from posixpath import split
from facebook_scraper import get_posts
import fbpost
import fbcomment
import fbgroup
import fbpage
import json
import time
import requests
from datetime import datetime
import openpyxl
import random
import es_conf as es

HEADERS = {"content-type": "application/json"}
# INPUT_FILE = "Lista grup do przeglądu.xlsx"
INPUT_FILE = "input/scraper_input.xlsx"


def get_iter_count():
    auth = requests.auth.HTTPBasicAuth(es.USER, es.PASSWORD)
    r = requests.get(f"{es.ES_URL}/{es.ITER_INDEX}/_doc/1", auth=auth)
    iter = r.json()["_source"]["iter_count"]
    new_iter = iter + 1 if iter < es.MAX_ITER else 1
    data = {"iter_count": new_iter}
    r = requests.put(
        f"{es.ES_URL}/{es.ITER_INDEX}/_doc/1",
        data=json.dumps(data),
        headers=HEADERS,
        auth=auth,
    )
    return iter


def to_epoch(date):
    utc_time = datetime.strptime(date, "%Y-%m-%d %H:%M:%S")
    return (utc_time - datetime(1970, 1, 1)).total_seconds()


def add_to_index(index, data):
    if index == es.POST_INDEX:
        auth = requests.auth.HTTPBasicAuth(es.USER, es.PASSWORD)
        r = requests.put(
            f"{es.ES_URL}/{index}/_doc/{data['post_id']}",
            headers=HEADERS,
            data=json.dumps(data),
            auth=auth,
        )
        # print(r.json())
    else:
        auth = requests.auth.HTTPBasicAuth(es.USER, es.PASSWORD)
        r = requests.put(
            f"{es.ES_URL}/{index}/_doc/{data['comment_id']}",
            headers=HEADERS,
            data=json.dumps(data),
            auth=auth,
        )
        # print(r.json())


def process_post(p, group=None, page=None):
    post = fbpost.Post(
        post_id=int(p["post_id"]),
        content=p["text"],
        date=to_epoch(str(p["time"])),
        post_url=p["post_url"],
        urls=p["link"],
        author_id=p["user_id"],
        author_profile_url=p["user_url"],
        like_count=p["likes"],
        comment_count=p["comments"],
        share_count=p["shares"],
    )

    if group:
        post.group_id = group.group_id
        post.group_name = group.group_name

    if page:
        post.page_id = page.page_id
        post.page_name = page.page_name

    return post


def process_comment(c, post_id, group=None, page=None):
    try:
        comment = fbcomment.Comment(
            original_post_id=post_id,
            comment_id=int(c["comment_id"]),
            content=c["comment_text"],
            date=to_epoch(str(c["comment_time"])),
            author_id=int(c["commenter_id"]),
            author_profile_url=c["commenter_url"],
            replies=str(c["replies"]),
            reaction_count=c["comment_reaction_count"],
        )
        if group:
            comment.group_id = group.group_id
            comment.group_name = group.group_name

        if page:
            comment.page_id = page.page_id
            comment.page_name = page.page_name
        return comment
    except ValueError:
        print(c)


def process_page(page):
    print(f"\nProcessing {page.page_name}, page_id: {page.page_id}")
    options = {"comments": True}
    
    page_count = random.randint(2,6)

    for p in get_posts(page.page_id, pages=page_count, options=options):
        print(".", end="")
        time.sleep(random.randint(5, 10))
        try:
            post = process_post(p=p, page=page)
        except Exception as e:
            print(e)
            print(p)
        for c in p["comments_full"]:
            print(".", end="")
            try:
                comment = process_comment(c, post_id=p["post_id"], page=page)
                if check_if_content_contains_selected_phrases(comment.content):
                    add_to_index(index=es.COMMENTS_INDEX, data=comment.to_dict())
            except TypeError:
                print(f"Error processing comment: {c}")
        if check_if_content_contains_selected_phrases(post.content):
            add_to_index(index=es.POST_INDEX, data=post.to_dict())  


def process_group(group):
    print(f"\nProcessing {group.group_name}, group id: {group.group_id}")
    options = {"comments": True}
    if group.private:
        return
    
    page_count = random.randint(2,6)
    for p in get_posts(group=group.group_id, pages=page_count, options=options):
        print(".", end="")
        time.sleep(random.randint(5, 10))
        try:
            post = process_post(p=p, group=group)
        except Exception as e:
            print(e)
            print(p)
        for c in p["comments_full"]:
            comment = process_comment(c, post_id=p["post_id"], group=group)
            if check_if_content_contains_selected_phrases(comment.content):
                add_to_index(index=es.COMMENTS_INDEX, data=comment.to_dict())
        if check_if_content_contains_selected_phrases(post.content):
            add_to_index(index=es.POST_INDEX, data=post.to_dict())


def get_jobs(iter):
    # open input file
    wb = openpyxl.load_workbook(INPUT_FILE)

    # get sheet for iter
    ws = wb[f"{iter}"]
    jobs = []

    # read groups and pages
    for row in range(1, ws.max_row + 1):
        try:
            # skip column headers
            if row == 1:
                continue

            # read group data
            if "groups" in ws[row][2].value:
                group_id = ws[row][2].value.split("/")[-2]
                group = fbgroup.Group(
                    group_id=group_id,
                    group_name=ws[row][1].value,
                    email="",
                    password="",
                    private=True if ws[row][3] == "1" else False,
                )
                jobs.append({"type": "group", "data": group})
            # read page data
            else:
                page_id = ws[row][2].value.split("/")[-2]
                page = fbpage.Page(page_id=page_id, page_name=ws[row][1].value)
                jobs.append({"type": "page", "data": page})

        except Exception as e:
            break
    return jobs


def print_jobs(jobs):
    for job in jobs:
        if job["type"] == "page":
            name = job["data"].page_name
            print(f"Page: {name}")
        elif job["type"] == "group":
            name = job["data"].group_name
            print(f"Group: {name}")
        else:
            print("Unknown job")


def get_phrases():
    phrases = []
    with open("phrases.txt", "r") as f:
        for line in f.readlines():
            phrases.append(line.replace("\n", "").lower())
    return phrases


def check_if_content_contains_selected_phrases(content):
    try:
        phrases = get_phrases()
        content_split = content.lower().split(" ")
        # return True
        return any(word in phrases for word in content_split)
    except AttributeError:
        # return True
        return False


def main():
    iter = get_iter_count()
    jobs = get_jobs(iter)
    random.shuffle(jobs)
    for job in jobs:
        time.sleep(random.randint(5,10) * 30)
        if job["type"] == "group":
            process_group(job["data"])
        elif job["type"] == "page":
            process_page(job["data"])
        else:
            print("Unknown type")


if __name__ == "__main__":
    main()
